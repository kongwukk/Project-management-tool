from __future__ import annotations

import html
import re
from dataclasses import dataclass
from datetime import datetime
from email import policy
from email.parser import BytesParser
from html.parser import HTMLParser
from pathlib import Path
from typing import Iterable, List

from .file_processor import FileProcessingError, parse_office_file
from .models import ParsedDocument


IMPORT_EXTENSIONS = {
    ".docx",
    ".xls",
    ".xlsx",
    ".xlsm",
    ".pptx",
    ".one",
    ".mht",
    ".mhtml",
    ".html",
    ".htm",
    ".txt",
    ".md",
}


class ProjectImportError(RuntimeError):
    pass


@dataclass
class ImportResult:
    output_path: Path
    documents: List[ParsedDocument]
    warnings: List[str]


def import_project_files(source_paths: Iterable[str | Path], output_path: str | Path, *, title: str = "") -> ImportResult:
    sources = [Path(path) for path in source_paths]
    if not sources:
        raise ProjectImportError("请选择至少一个要导入的旧项目文件。")

    documents: List[ParsedDocument] = []
    warnings: List[str] = []
    for source in sources:
        try:
            documents.append(parse_import_source(source))
        except Exception as exc:
            warnings.append(f"{source.name}: {exc}")

    if not documents:
        detail = "\n".join(warnings) if warnings else "没有可导入的内容。"
        raise ProjectImportError(f"导入失败：{detail}")

    target = Path(output_path)
    if target.suffix.lower() != ".md":
        target = target.with_suffix(".md")
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(render_project_markdown(documents, title=title or target.stem), encoding="utf-8")
    return ImportResult(output_path=target, documents=documents, warnings=warnings)


def parse_import_source(path: str | Path) -> ParsedDocument:
    source = Path(path)
    suffix = source.suffix.lower()
    if suffix == ".xls":
        return ParsedDocument(source.name, "Excel", _parse_xls(source), str(source))
    if suffix == ".xlsm":
        return ParsedDocument(source.name, "Excel", _parse_xlsm(source), str(source))
    if suffix in {".docx", ".xlsx", ".pptx"}:
        return parse_office_file(source)
    if suffix in {".mht", ".mhtml"}:
        return ParsedDocument(source.name, "OneNote/MHTML", _parse_mhtml(source), str(source))
    if suffix in {".html", ".htm"}:
        return ParsedDocument(source.name, "OneNote/HTML", _parse_html(source), str(source))
    if suffix == ".one":
        return ParsedDocument(source.name, "OneNote", _parse_onenote_binary(source), str(source))
    if suffix in {".txt", ".md"}:
        return ParsedDocument(source.name, "Text", source.read_text(encoding="utf-8", errors="replace"), str(source))
    raise FileProcessingError(
        f"不支持的导入格式：{suffix or '未知'}。支持 .docx、.xls、.xlsx、.xlsm、.one、.mht、.html、.txt、.md。"
    )


def render_project_markdown(documents: List[ParsedDocument], *, title: str) -> str:
    now = datetime.now().isoformat(timespec="seconds")
    lines: List[str] = [
        f"# {title}",
        "",
        "## 项目概览",
        "",
        f"- 创建时间：{now}",
        "- 来源：旧项目管理文件导入",
        f"- 导入文件数：{len(documents)}",
        "",
        "## 当前状态",
        "",
        "- [ ] 请根据导入内容梳理项目目标、范围和下一步行动。",
        "",
        "## 待办事项",
        "",
        "- [ ] 检查导入内容并补充负责人、截止时间和优先级。",
        "",
        "## 导入来源",
        "",
    ]

    for index, document in enumerate(documents, start=1):
        lines.extend(
            [
                f"### {index}. {document.filename}",
                "",
                f"- 类型：{document.source_type}",
                f"- 路径：{document.path or '未记录'}",
                "",
                "#### 内容",
                "",
                _normalize_markdown_content(document.content),
                "",
            ]
        )

    return "\n".join(lines).rstrip() + "\n"


def _parse_xlsm(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise FileProcessingError("缺少 openpyxl 依赖，请先运行：pip install openpyxl") from exc

    workbook = load_workbook(filename=str(path), data_only=True, read_only=True, keep_vba=False)
    sections: List[str] = [f"# 文件：{path.name}"]
    for sheet in workbook.worksheets:
        sections.append(f"## 工作表：{sheet.title}")
        rows_written = 0
        for row in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value) for value in row]
            if not any(value.strip() for value in values):
                continue
            sections.append(", ".join(values))
            rows_written += 1
        if rows_written == 0:
            sections.append("（空工作表）")
    workbook.close()
    return "\n".join(sections)


def _parse_xls(path: Path) -> str:
    try:
        import pandas as pd
    except ImportError as exc:
        raise FileProcessingError("缺少 pandas/xlrd 依赖，请先运行：pip install pandas xlrd") from exc

    try:
        sheets = pd.read_excel(path, sheet_name=None, dtype=str)
    except ImportError as exc:
        raise FileProcessingError("缺少 xlrd 依赖，请先运行：pip install xlrd") from exc

    sections: List[str] = [f"# 文件：{path.name}"]
    for sheet_name, frame in sheets.items():
        sections.append(f"## 工作表：{sheet_name}")
        frame = frame.fillna("")
        if frame.empty:
            sections.append("（空工作表）")
        else:
            sections.append(frame.to_csv(index=False).strip())
    return "\n".join(sections)


def _parse_mhtml(path: Path) -> str:
    with path.open("rb") as handle:
        message = BytesParser(policy=policy.default).parse(handle)

    parts: List[str] = []
    if message.is_multipart():
        for part in message.walk():
            content_type = part.get_content_type()
            if content_type not in {"text/html", "text/plain"}:
                continue
            payload = part.get_content()
            parts.append(_html_to_text(payload) if content_type == "text/html" else str(payload))
    else:
        payload = message.get_content()
        parts.append(_html_to_text(payload) if message.get_content_type() == "text/html" else str(payload))

    text = "\n\n".join(part.strip() for part in parts if part and part.strip())
    return text or "（未从 MHTML 文件中提取到文本）"


def _parse_html(path: Path) -> str:
    raw = path.read_text(encoding="utf-8", errors="replace")
    return _html_to_text(raw)


def _parse_onenote_binary(path: Path) -> str:
    data = path.read_bytes()
    strings = set()
    for pattern in (rb"(?:[\x20-\x7e]\x00){4,}", rb"[\x20-\x7e]{4,}"):
        for match in re.finditer(pattern, data):
            raw = match.group(0)
            encoding = "utf-16le" if b"\x00" in raw else "utf-8"
            try:
                text = raw.decode(encoding, errors="ignore")
            except Exception:
                continue
            text = _clean_extracted_text(text)
            if _looks_useful(text):
                strings.add(text)

    if not strings:
        return (
            "（未能从 OneNote .one 文件中提取到可读文本。"
            "建议在 OneNote 中将页面导出为 .docx、.mht 或 .html 后再导入。）"
        )
    return "\n".join(sorted(strings))


def _html_to_text(raw: str) -> str:
    parser = _TextHTMLParser()
    parser.feed(raw)
    return parser.text()


class _TextHTMLParser(HTMLParser):
    BLOCK_TAGS = {"p", "div", "br", "li", "tr", "table", "section", "article", "h1", "h2", "h3", "h4", "h5", "h6"}

    def __init__(self) -> None:
        super().__init__()
        self._parts: List[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() in self.BLOCK_TAGS:
            self._parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in self.BLOCK_TAGS:
            self._parts.append("\n")

    def handle_data(self, data: str) -> None:
        text = html.unescape(data).strip()
        if text:
            self._parts.append(text)

    def text(self) -> str:
        joined = " ".join(self._parts)
        lines = [" ".join(line.split()) for line in joined.splitlines()]
        return "\n".join(line for line in lines if line).strip()


def _normalize_markdown_content(content: str) -> str:
    text = content.strip()
    if not text:
        return "（未提取到内容）"
    return "\n".join(line.rstrip() for line in text.splitlines()).strip()


def _clean_extracted_text(text: str) -> str:
    text = text.replace("\x00", "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _looks_useful(text: str) -> bool:
    if len(text) < 4:
        return False
    if len(set(text)) <= 2:
        return False
    return any(char.isalnum() or "\u4e00" <= char <= "\u9fff" for char in text)
